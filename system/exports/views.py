from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from system.users.decorators import role_required
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_GET, require_POST
from .models import ExportRequest, can_export_direct, must_request_export
from system.users.models import User
from shared.projects.models import Project
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from io import BytesIO


@login_required
@role_required(allowed_roles=["UESO", "VP", "DIRECTOR"])
def exports_view(request):
    requests = ExportRequest.objects.all()

    # Filters
    sort_by = request.GET.get('sort_by', 'date_submitted')
    order = request.GET.get('order', 'desc')
    status = request.GET.get('status', '')
    date = request.GET.get('date', '')
    search = request.GET.get('search', '').strip()

    # Apply filters
    if status: requests = requests.filter(status__iexact=status)
    if date: requests = requests.filter(deadline__date=date)
    if search: requests = requests.filter(submitted_by__icontains=search)

    requests = requests.distinct()

    # Sorting
    sort_map = {
        'date_submitted': 'date_submitted',
        'status': 'status',
        'type': 'type',
    }
    sort_field = sort_map.get(sort_by, 'date_submitted')
    if sort_field:
        if order == 'desc':
            sort_field = '-' + sort_field
        requests = requests.order_by(sort_field)

    # Filter Options
    all_statuses = [status[1] for status in ExportRequest._meta.get_field('status').choices]

    # Pagination
    paginator = Paginator(requests, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    page_range = paginator.get_elided_page_range(page_obj.number)

    return render(request, 'exports/exports.html', {
        'search': search,
        'sort_by': sort_by,
        'order': order,
        'all_statuses': all_statuses,
        'status': status,
        'date': date,
        'page_obj': page_obj,
        'page_range': page_range,
        'querystring': request.GET.urlencode().replace('&page='+str(page_obj.number), '') if page_obj else '',
    })



@require_GET
@login_required
@role_required(allowed_roles=["UESO", "VP", "DIRECTOR"])
def export_manage_user(request):
    user = request.user
    UserModel = User
    users = UserModel.objects.all()
    query_params = {}

    # --- Filters (match manage_user view) ---
    from django.db.models import Q
    search = request.GET.get('search', '').strip()
    if search:
        users = users.filter(
            Q(given_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(middle_initial__icontains=search) |
            Q(suffix__icontains=search) |
            Q(email__icontains=search)
        )
        query_params['search'] = search

    sort_by = request.GET.get('sort_by', 'date')
    order = request.GET.get('order', 'desc')
    role = request.GET.get('role', '')
    verified = request.GET.get('verified', '')
    date = request.GET.get('date', '')
    college = request.GET.get('college', '')
    campus = request.GET.get('campus', '')

    if sort_by:
        query_params['sort_by'] = sort_by
    if order:
        query_params['order'] = order
    if role:
        users = users.filter(role=role)
        query_params['role'] = role
    if verified == 'true':
        users = users.filter(is_confirmed=True)
        query_params['verified'] = 'true'
    elif verified == 'false':
        users = users.filter(is_confirmed=False)
        query_params['verified'] = 'false'
    if date:
        users = users.filter(date_joined__date=date)
        query_params['date'] = date
    if college:
        users = users.filter(college_id=college)
        query_params['college'] = college
    if campus:
        users = users.filter(campus=campus)
        query_params['campus'] = campus

    # Sorting
    if sort_by == 'name':
        sort_field = ['last_name', 'given_name', 'middle_initial', 'suffix']
    else:
        sort_map = {
            'email': 'email',
            'date': 'date_joined',
            'role': 'role',
        }
        sort_field = [sort_map.get(sort_by, 'last_name')]
    if order == 'desc':
        sort_field = ['-' + f for f in sort_field]
    users = users.order_by(*sort_field)

    # Generate XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manage Users"
    headers = [
        'Last Name', 'Given Name', 'Middle Initial', 'Suffix', 'Email', 'Role', 'Verified', 'Date Joined', 'College', 'Campus'
    ]
    ws.append(headers)
    for u in users:
        ws.append([
            u.last_name,
            u.given_name,
            u.middle_initial,
            u.suffix,
            u.email,
            u.get_role_display() if hasattr(u, 'get_role_display') else getattr(u, 'role', ''),
            u.is_confirmed,
            u.date_joined.strftime('%Y-%m-%d %H:%M'),
            str(getattr(u, 'college', '')),
            u.get_campus_display() if hasattr(u, 'get_campus_display') else getattr(u, 'campus', ''),
        ])
    # Auto-fit column widths
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="manage_users_export.xlsx"'
    return response


@require_GET
@login_required
def export_project(request):
    user = request.user
    from django.db.models import Q
    projects = Project.objects.all()
    # Filters (match admin_project view)
    search = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort_by', 'last_updated')
    order = request.GET.get('order', 'desc')
    college = request.GET.get('college', '')
    campus = request.GET.get('campus', '')
    agenda = request.GET.get('agenda', '')
    status = request.GET.get('status', '')
    year = request.GET.get('year', '')
    quarter = request.GET.get('quarter', '')
    date = request.GET.get('date', '')

    # Filter by college/campus via team leader
    if college:
        projects = projects.filter(project_leader__college__id=college)
    if campus:
        projects = projects.filter(project_leader__campus=campus)
    if agenda:
        projects = projects.filter(agenda__id=agenda)
    if status:
        projects = projects.filter(status=status)
    if year:
        projects = projects.filter(start_date__year=year)
    if quarter:
        # Q1: Jan-Mar, Q2: Apr-Jun, Q3: Jul-Sep, Q4: Oct-Dec
        qmap = {'1': (1,3), '2': (4,6), '3': (7,9), '4': (10,12)}
        if quarter in qmap:
            start, end = qmap[quarter]
            projects = projects.filter(start_date__month__gte=start, start_date__month__lte=end)
    if date:
        projects = projects.filter(start_date=date)
    if search:
        projects = projects.filter(title__icontains=search)

    # Sorting
    sort_map = {
        'title': 'title',
        'last_updated': 'updated_at',
        'start_date': 'start_date',
        'progress': '', # Placeholder, not supported in DB sort
    }
    sort_field = sort_map.get(sort_by, 'title')
    if sort_field:
        if order == 'desc':
            sort_field = '-' + sort_field
        projects = projects.order_by(sort_field)
    # If progress sort, sort in Python (not supported in DB)
    elif sort_by == 'progress':
        projects = sorted(projects, key=lambda p: (p.progress[0] / p.progress[1]) if p.progress[1] else 0, reverse=(order=='desc'))

    if can_export_direct(user):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Alignment
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        headers = [
            'Title', 'Leader', 'College/Unit', 'Last Updated', 'Start Date', 'Progress', 'Status'
        ]
        ws.append(headers)
        for p in projects:
            ws.append([
                p.title,
                p.project_leader.get_full_name() if p.project_leader else '',
                p.project_leader.college.name if p.project_leader and p.project_leader.college else '',
                p.updated_at.strftime('%Y-%m-%d') if p.updated_at else '',
                p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                getattr(p, 'progress_display', ''),
                p.get_status_display() if hasattr(p, 'get_status_display') else p.status,
            ])
        # Auto-fit column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="projects_export.xlsx"'
        return response
    elif must_request_export(user):
        ExportRequest.objects.create(
            type='PROJECT',
            date_submitted=timezone.now(),
            submitted_by=user,
            status='PENDING',
        )
        return JsonResponse({'message': 'Your export request has been submitted for approval.'}, status=202)
    else:
        return JsonResponse({'error': 'You do not have permission to export.'}, status=403)


@require_GET
@login_required
def export_log(request):
    # Later
    return 0


@require_GET
@login_required
def export_goals(request):
    # Later
    return 0


@require_GET
@login_required
def export_budget(request):
    # Later
    return 0